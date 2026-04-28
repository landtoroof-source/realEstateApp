from fastapi import FastAPI, HTTPException, File, UploadFile, Header, Header
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from pydantic import BaseModel, EmailStr
from typing import Optional
import os
from groq import Groq
from datetime import datetime
import json
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import shutil

app = FastAPI(title="RealEstate AI EdTech", version="1.0.0")

# app.add_middleware(
#     CORSMiddleware,
#     allow_origins=["*"],
#     allow_credentials=True,
#     allow_methods=["*"],
#     allow_headers=["*"],
# )
app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "https://realestate-frontend-mu.vercel.app",
        "http://localhost:3000",
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

print("Loading environment variables...")
load_dotenv()  # Load from .env file
# print("Environment variables loaded:")
# print(f"GROQ_API_KEY: {os.environ.get('GROQ_API_KEY')}")
# print(f"ADMIN_SECRET: {os.environ.get('ADMIN_SECRET')}")
# ── Groq client ──────────────────────────────────────────────────────────────
client = Groq(api_key=os.environ.get("GROQ_API_KEY"))
MODEL = "llama-3.1-8b-instant"

# ── File-based lead store ───────────────────────────────────────────────────────
LEADS_JSON_FILE = "leads.json"
LEADS_EXCEL_FILE = "leads.xlsx"

def load_leads() -> list[dict]:
    """Load leads from JSON file."""
    if os.path.exists(LEADS_JSON_FILE):
        try:
            with open(LEADS_JSON_FILE, "r") as f:
                return json.load(f)
        except (json.JSONDecodeError, IOError):
            return []
    return []

def save_leads(leads: list[dict]) -> None:
    """Save leads to JSON file."""
    with open(LEADS_JSON_FILE, "w") as f:
        json.dump(leads, f, indent=2)

def export_to_excel(leads: list[dict], filename: str = LEADS_EXCEL_FILE) -> None:
    """Export leads to Excel file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"
    
    # Add headers
    headers = ["Session ID", "First Name", "Last Name", "Email", "Phone", "Budget", "Purpose", "Location", "Timestamp", "Source"]
    ws.append(headers)
    
    # Add data
    for lead in leads:
        row = [
            lead.get("session_id", ""),
            lead.get("first_name", ""),
            lead.get("last_name", ""),
            lead.get("email", ""),
            lead.get("phone", ""),
            lead.get("budget", ""),
            lead.get("purpose", ""),
            lead.get("location", ""),
            lead.get("timestamp", ""),
            lead.get("source", "")
        ]
        ws.append(row)
    
    # Adjust column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20
    
    wb.save(filename)

def import_from_excel(filename: str) -> list[dict]:
    """Import leads from Excel file."""
    from openpyxl import load_workbook
    
    leads = []
    try:
        wb = load_workbook(filename)
        ws = wb.active
        
        # Skip header row
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:  # Check if session_id exists
                lead = {
                    "session_id": row[0],
                    "first_name": row[1] or "",
                    "last_name": row[2] or "",
                    "email": row[3] or "",
                    "phone": row[4] or "",
                    "budget": row[5] or "",
                    "purpose": row[6] or "",
                    "location": row[7] or "",
                    "timestamp": row[8] or "",
                    "source": row[9] or "imported_excel"
                }
                leads.append(lead)
    except Exception as e:
        print(f"Error importing Excel file: {e}")
    
    return leads

# ── System prompt (paste your own below) ─────────────────────────────────────
SYSTEM_PROMPT = """
You are an expert AI real estate advisor and educator working for a premium EdTech platform.

Your goals:
1. Educate users about real estate — buying, renting, investing, market trends, and property valuation.
2. Give personalized, actionable suggestions based on the user's budget, location, and goals.
3. Build trust by explaining concepts clearly (EMI, ROI, RERA, carpet area, etc.).
4. Naturally guide the conversation toward capturing the user's name, email, and phone number
   so our team can follow up with exclusive deals and free consultations.

Lead capture rules:
- After 2–3 helpful exchanges, politely ask: "Would you like a FREE consultation call with our
  real estate experts? Just share your name and contact details."
- If user shares contact info, confirm you've noted it and say the team will reach out within 24hrs.
- Never be pushy. Be warm, knowledgeable, and helpful first.

Always respond in the same language the user writes in (English or local language mix is fine).
Keep responses concise — 3–5 sentences max unless explaining a complex topic.
"""

# ── Schemas ───────────────────────────────────────────────────────────────────
class ChatMessage(BaseModel):
    session_id: str
    message: str
    history: Optional[list[dict]] = []

class Lead(BaseModel):
    session_id: str
    first_name: str
    last_name: str
    email: Optional[str] = None
    phone: Optional[str] = None
    budget: Optional[str] = None
    purpose: Optional[str] = None
    location: Optional[str] = None

class ChatResponse(BaseModel):
    reply: str
    session_id: str
    lead_detected: bool = False

# ── Authorization helper ──────────────────────────────────────────────────────
def verify_admin_secret(authorization: str = Header(None)) -> None:
    """Verify admin secret from Authorization header."""
    if not authorization:
        raise HTTPException(status_code=401, detail="Missing Authorization header")
    
    # Expected format: "Bearer <secret>"
    parts = authorization.split()
    if len(parts) != 2 or parts[0].lower() != "bearer":
        raise HTTPException(status_code=401, detail="Invalid Authorization header format. Use: Bearer <secret>")
    
    secret = parts[1]
    if secret != os.environ.get("ADMIN_SECRET", "changeme"):
        raise HTTPException(status_code=403, detail="Invalid secret")

# ── Simple keyword-based lead detection ───────────────────────────────────────
def detect_lead_info(text: str) -> dict:
    import re
    lead = {}
    email = re.findall(r'[\w.+-]+@[\w-]+\.\w+', text)
    phone = re.findall(r'[\+]?[\d\s\-]{10,13}', text)
    if email:
        lead["email"] = email[0]
    if phone:
        lead["phone"] = phone[0].strip()
    return lead

# ── Routes ────────────────────────────────────────────────────────────────────
@app.get("/")
def root():
    return {"status": "ok", "service": "RealEstate AI EdTech API"}

@app.post("/chat", response_model=ChatResponse)
async def chat(payload: ChatMessage):
    try:
        messages = [{"role": "system", "content": SYSTEM_PROMPT}]

        # Include conversation history
        for h in payload.history[-10:]:  # keep last 10 turns
            messages.append(h)

        messages.append({"role": "user", "content": payload.message})

        response = client.chat.completions.create(
            model=MODEL,
            messages=messages,
            temperature=0.7,
            max_tokens=512,
        )

        reply = response.choices[0].message.content

        # Check if user shared contact info
        lead_info = detect_lead_info(payload.message)
        lead_detected = bool(lead_info)

        if lead_detected:
            leads = load_leads()
            leads.append({
                "session_id": payload.session_id,
                "timestamp": datetime.utcnow().isoformat(),
                **lead_info
            })
            save_leads(leads)

        return ChatResponse(
            reply=reply,
            session_id=payload.session_id,
            lead_detected=lead_detected
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/leads/capture")
def capture_lead(lead: Lead):
    """Explicit lead capture endpoint (call when user fills a form)."""
    record = {
        **lead.dict(),
        "timestamp": datetime.utcnow().isoformat(),
        "source": "explicit_form"
    }
    leads = load_leads()
    leads.append(record)
    save_leads(leads)
    return {"status": "captured", "lead_id": len(leads)}


@app.get("/leads")
def get_leads(authorization: str = Header(None)):
    """Simple admin endpoint — protect with secret key in prod."""
    verify_admin_secret(authorization)
    leads = load_leads()
    return {"total": len(leads), "leads": leads}

@app.get("/leads/export/excel")
def export_leads_excel(authorization: str = Header(None)):
    """Export leads to Excel file."""
    verify_admin_secret(authorization)
    
    leads = load_leads()
    export_to_excel(leads)
    
    return FileResponse(
        path=LEADS_EXCEL_FILE,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=LEADS_EXCEL_FILE
    )

@app.post("/leads/import/excel")
async def import_leads_excel(file: UploadFile = File(...), authorization: str = Header(None)):
    """Import leads from Excel file."""
    verify_admin_secret(authorization)
    
    try:
        # Save uploaded file temporarily
        temp_file = f"temp_{file.filename}"
        with open(temp_file, "wb") as f:
            shutil.copyfileobj(file.file, f)
        
        # Import leads from the file
        imported_leads = import_from_excel(temp_file)
        
        # Get existing leads and append new ones
        existing_leads = load_leads()
        existing_leads.extend(imported_leads)
        save_leads(existing_leads)
        
        # Clean up temp file
        os.remove(temp_file)
        
        return {
            "status": "imported",
            "count": len(imported_leads),
            "total": len(existing_leads)
        }
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.get("/health")
def health():
    return {"status": "healthy", "model": MODEL, "timestamp": datetime.utcnow().isoformat()}
